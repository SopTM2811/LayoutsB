from fastapi import APIRouter, Request, Form, Depends, HTTPException, Query, Body
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_
from io import StringIO, BytesIO
from openpyxl import load_workbook
from pathlib import Path

import csv
import re
import unicodedata
import zipfile

from io import BytesIO
from app.db import get_db
from app.models import (
    CatBeneficiario,
    CatCuentaBeneficiario,
    CatInstitucionSTP,
    CatInstitucionBancaria,
    CatLayout,
    CatLayoutCampo,
    CatOrigenOperativo,
    CatOrigenOpcion,
)

MAPA_BANCOS = {}
router = APIRouter(tags=["Operaciones"])
templates = Jinja2Templates(directory="app/templates")


# =========================
# FORM INICIAL PARA SELECCIÓN DE LAYOUT
# =========================
def limpiar_texto(texto, tamaño):

    if not texto:
        texto = ""

    # Quitar acentos
    texto = unicodedata.normalize("NFD", texto)
    texto = texto.encode("ascii", "ignore").decode("utf-8")

    # Reemplazar símbolos por espacio
    texto = re.sub(r"[^A-Za-z0-9 ]", " ", texto)

    texto = texto.upper()

    # Cortar
    texto = texto[:tamaño]

    # Rellenar espacios derecha
    texto = texto.ljust(tamaño)

    return texto


def cargar_bancos():
    global MAPA_BANCOS

    if MAPA_BANCOS:
        return

    BASE_DIR = Path(__file__).resolve().parent.parent
    ruta = (
        BASE_DIR
        / "plantillas_layouts"
        / "ESPIRAL"
        / "TRANSFERENCIAS_FormatoDisMasiva20Oct2025.xlsx"
    )

    if not ruta.exists():
        raise Exception(f"No se encontró la plantilla: {ruta}")

    wb = load_workbook(ruta, data_only=True)

    ws = wb["Bancos"]

    for row in ws.iter_rows(min_row=2, values_only=True):

        prefijo = str(row[0]).strip()
        nombre = row[1]
        codigo = row[2]

        MAPA_BANCOS[prefijo] = {"nombre": nombre, "codigo": codigo}

    print("Bancos cargados:", len(MAPA_BANCOS))


def obtener_banco_stp(prefijo: str, db: Session):

    banco = (
        db.query(CatInstitucionSTP)
        .filter(CatInstitucionSTP.prefijo_clabe == prefijo)
        .first()
    )

    if banco:
        return banco.participante

    return ""


def obtener_banco_santander(prefijo: str, db: Session):
    # Los primeros 3 dígitos de CLABE = clave banxico (ej: "012")
    banco = (
        db.query(CatInstitucionBancaria)
        .filter(CatInstitucionBancaria.clave_banxico == prefijo)
        .first()
    )
    if banco:
        return banco.clave_transfer  # ej: "BACOM"
    return ""


@router.get("/", response_class=HTMLResponse)
async def form_layout(request: Request, db: Session = Depends(get_db)):

    origenes = (
        db.query(CatOrigenOperativo).filter(CatOrigenOperativo.activo == True).all()
    )

    beneficiarios = (
        db.query(CatBeneficiario).filter(CatBeneficiario.activo == True).all()
    )

    return templates.TemplateResponse(
        "operaciones/form.html",
        {
            "request": request,
            "origenes": origenes,
            "beneficiarios": beneficiarios,  # 🔥 IMPORTANTE
        },
    )


# =========================
# OBTENER OPCIONES POR ORIGEN
# =========================
@router.get("/api/opciones/{id_origen}")
async def opciones_por_origen(id_origen: int, db: Session = Depends(get_db)):
    opciones = (
        db.query(CatOrigenOpcion)
        .filter(CatOrigenOpcion.id_origen == id_origen, CatOrigenOpcion.activo == True)
        .all()
    )
    return JSONResponse(
        [{"id_opcion": o.id_opcion, "codigo": o.codigo} for o in opciones]
    )


# =========================
# OBTENER LAYOUT Y CAMPOS POR OPCIÓN
# =========================
@router.get("/api/layout/{id_opcion}")
async def layout_por_opcion(id_opcion: int, db: Session = Depends(get_db)):

    layout = (
        db.query(CatLayout)
        .filter(CatLayout.id_opcion == id_opcion, CatLayout.activo == True)
        .first()
    )

    if not layout:
        raise HTTPException(status_code=404, detail="Layout no encontrado")

    campos = (
        db.query(CatLayoutCampo)
        .filter(
            CatLayoutCampo.id_layout == layout.id_layout, CatLayoutCampo.activo == True
        )
        .order_by(CatLayoutCampo.posicion_inicio)
        .all()
    )

    return {
        "id_layout": layout.id_layout,
        "codigo": layout.codigo,
        "tipo_layout": layout.tipo_layout,
        "header": [
            {
                "nombre_logico": c.nombre_logico,
                "etiqueta_ui": c.etiqueta_ui,
            }
            for c in campos
            if c.tipo_campo == "HEADER"
        ],
        "detalle": [
            {
                "nombre_logico": c.nombre_logico,
                "etiqueta_ui": c.etiqueta_ui,
                "editable": c.editable,
                "valor_fijo": c.valor_fijo,
            }
            for c in campos
            if c.tipo_campo == "DETALLE_UI"
        ],
    }


@router.post("/api/buscar-cuentas")
async def buscar_cuentas(data: dict, db: Session = Depends(get_db)):

    registros = data.get("registros", [])

    cuentas = [r["cuenta"] for r in registros]

    resultados = (
        db.query(CatCuentaBeneficiario)
        .join(CatBeneficiario)
        .filter(CatCuentaBeneficiario.numero_cuenta.in_(cuentas))
        .all()
    )

    encontrados = []
    encontrados_cuentas = []

    for c in resultados:
        reg_original = next(r for r in registros if r["cuenta"] == c.numero_cuenta)

        encontrados.append(
            {
                "beneficiario": c.beneficiario.nombre,
                "clabe": c.numero_cuenta,
                "importe": reg_original.get("importe", ""),
                "concepto": reg_original.get("concepto", ""),
            }
        )

        encontrados_cuentas.append(c.numero_cuenta)

    # detectar cuentas no registradas
    no_encontrados = []
    for r in registros:

        if r["cuenta"] not in encontrados_cuentas:

            no_encontrados.append(
                {
                    "cuenta": r.get("cuenta"),
                    "nombre": r.get("nombre", ""),
                    "importe": r.get("importe", ""),
                    "concepto": r.get("concepto", ""),
                }
            )

    return {"encontrados": encontrados, "no_encontrados": no_encontrados}


@router.get("/api/buscar-beneficiario")
async def buscar_beneficiario(
    q: str = Query(...), id_layout: int = Query(...), db: Session = Depends(get_db)
):

    if not q:
        return []

    resultados = (
        db.query(CatCuentaBeneficiario)
        .join(CatBeneficiario)
        .filter(
            CatCuentaBeneficiario.activo == True,
            CatBeneficiario.activo == True,
            or_(
                CatCuentaBeneficiario.numero_cuenta.ilike(f"%{q}%"),
                CatBeneficiario.nombre.ilike(f"%{q}%"),
            ),
        )
        .all()
    )

    return [
        {
            "id": c.beneficiario.id_beneficiario,
            "nombre": f"{c.numero_cuenta} - {c.beneficiario.nombre}",
            # 🔥 Datos completos para autocompletar
            "beneficiario": c.beneficiario.nombre,
            # "rfc": c.beneficiario.rfc,
            "clabe": c.numero_cuenta,
            "clabe_o_tarjeta": c.numero_cuenta,
        }
        for c in resultados
    ]


@router.get("/api/banco")
async def obtener_banco(prefijo: str):

    banco = MAPA_BANCOS.get(prefijo)

    if banco:
        return banco

    return {"nombre": "", "codigo": ""}


@router.get("/api/banco-stp")
async def obtener_banco_stp_api(prefijo: str, db: Session = Depends(get_db)):

    banco = (
        db.query(CatInstitucionSTP)
        .filter(CatInstitucionSTP.prefijo_clabe == prefijo)
        .first()
    )

    if banco:
        return {"codigo": banco.participante}

    return {"codigo": ""}


@router.get("/api/banco-santander")
async def obtener_banco_santander_api(prefijo: str, db: Session = Depends(get_db)):
    banco = (
        db.query(CatInstitucionBancaria)
        .filter(CatInstitucionBancaria.clave_banxico == prefijo)
        .first()
    )
    if banco:
        return {
            "nombre": banco.nombre_institucion,
            "clave_transfer": banco.clave_transfer,
        }
    return {"nombre": "", "clave_transfer": ""}


@router.post("/generar")
async def generar_archivo(data: dict = Body(...), db: Session = Depends(get_db)):

    header_data = data.get("header", {})
    id_layout = data.get("id_layout")
    registros = data.get("registros", [])

    if not registros:
        raise HTTPException(
            status_code=400, detail="No hay registros para generar archivo"
        )

    layout = db.get(CatLayout, id_layout)

    if not layout:
        raise HTTPException(status_code=404, detail="Layout no encontrado")

    if not layout.tipo_layout:
        raise HTTPException(
            status_code=400, detail="Layout sin tipo_layout configurado"
        )

    tipo = layout.tipo_layout.upper()

    # ======================================================
    # 🔵 DELIMITADO (CSV / MIFEL / KUSPIT)
    # ======================================================
    if tipo == "DELIMITADO":
        separador = layout.separador or ","

        if separador == "\\t":
            separador = "\t"

        codigo_layout = layout.codigo.lower()
        es_stp = "stp layout" in codigo_layout
        es_mifel = "mifel" in codigo_layout
        es_kuspit = "kuspit layout" in codigo_layout

        campos = (
            db.query(CatLayoutCampo)
            .filter(
                CatLayoutCampo.id_layout == id_layout, CatLayoutCampo.activo == True
            )
            .order_by(CatLayoutCampo.posicion_inicio)
            .all()
        )

        header_campos = [c for c in campos if c.tipo_campo in ("HEADER", "HEADER_TXT")]
        detalle_campos = sorted(
            [c for c in campos if c.tipo_campo in ("DETALLE_UI", "DETALLE_TXT")],
            key=lambda x: x.posicion_inicio,
        )

        lineas = []
        # ENCABEZADOS (solo para KUSPIT)
        if es_kuspit:
            encabezados = [c.etiqueta_ui for c in header_campos + detalle_campos]
            lineas.append(separador.join(encabezados))
        # ENCABEZADOS (solo para STP)
        if es_stp:
            encabezados = [c.etiqueta_ui for c in detalle_campos]
            lineas.append(separador.join(encabezados))

        suma_importes = sum(float(r.get("importe", 0) or 0) for r in registros)

        if not es_stp and not es_mifel:
            fila_header = []
            for campo in header_campos:
                if campo.nombre_logico == "ingreso_importe":
                    valor = suma_importes
                else:
                    valor = header_data.get(campo.nombre_logico, "")
                fila_header.append(str(valor))
            # completar columnas detalle vacías
            fila_header += [""] * len(detalle_campos)
            lineas.append(separador.join(fila_header))

        if es_mifel:
            header_txt_campos = sorted(
                [c for c in campos if c.tipo_campo in ("HEADER", "HEADER_TXT")],
                key=lambda x: x.posicion_inicio,
            )
            fila_header_mifel = []
            for campo in header_txt_campos:
                if campo.nombre_logico == "total_registros":
                    valor = str(len(registros))
                elif campo.nombre_logico == "total_importe":
                    valor = f"{suma_importes:.2f}"
                elif campo.valor_fijo is not None:
                    valor = str(campo.valor_fijo)
                else:
                    valor = str(header_data.get(campo.nombre_logico, "") or "")
                fila_header_mifel.append(valor)

            linea_header = "|" + "|".join(fila_header_mifel) + ";"
            lineas.append(linea_header)
        # ======================================================
        # FILAS DETALLE
        # ======================================================
        for i, registro in enumerate(registros, start=1):
            if es_mifel:
                fila = []
                for campo in detalle_campos:
                    if campo.nombre_logico == "consecutivo":
                        valor = str(i)
                    elif campo.valor_fijo is not None:
                        valor = str(campo.valor_fijo)
                    elif campo.nombre_logico in ["concepto", "descripcion"]:
                        valor = registro.get(campo.nombre_logico, "") or registro.get(
                            "concepto", ""
                        )
                    else:
                        valor = str(registro.get(campo.nombre_logico, "") or "")
                    fila.append(valor)

                linea = "|" + "|".join(fila) + ";"
                lineas.append(linea)
                continue

            fila = []

            if not es_stp and not es_mifel:
                for campo in header_campos:
                    if campo.nombre_logico == "esuelv":
                        fila.append(header_data.get("esuelv", ""))
                    else:
                        fila.append("")

            for campo in detalle_campos:
                # STP → obtener institución por CLABE
                nombre = campo.nombre_logico.strip().lower()

                if nombre == "institucion_contraparte":
                    cuenta = registro.get("cuenta_beneficiario", "") or registro.get(
                        "clabe", ""
                    )
                    prefijo = str(cuenta).replace("´'", "")[:3]
                    valor = obtener_banco_stp(prefijo, db) or ""
                # VALOR FIJO
                elif campo.valor_fijo is not None:
                    valor = campo.valor_fijo
                # VALOR DINAMICO
                else:
                    valor = registro.get(campo.nombre_logico, "")

                if valor is None:
                    valor == ""

                if campo.nombre_logico in ["clabe"] and not es_stp:
                    valor = "'" + str(valor)

                fila.append(str(valor))

                # ✅ VALIDACIÓN STP — CLABE debe tener exactamente 18 dígitos
            if es_stp:
                clabe_val = (
                    registro.get("clabe") or registro.get("cuenta_beneficiario") or ""
                )
                clabe_val = str(clabe_val).replace("'", "").strip()
                if len(clabe_val) != 18 or not clabe_val.isdigit():
                    beneficiario = registro.get("beneficiario", f"registro {i}")
                    raise HTTPException(
                        status_code=400,
                        detail=f"CLABE inválida para '{beneficiario}': '{clabe_val}' — debe tener exactamente 18 dígitos numéricos.",
                    )

            lineas.append(separador.join(fila))

        contenido = "\n".join(lineas)

        extension = (layout.tipo_archivo or "csv").lower()
        media = "text/plain" if extension == "txt" else "text/csv"

        return StreamingResponse(
            BytesIO(contenido.encode("utf-8")),
            media_type=media,
            headers={
                "Content-Disposition": f"attachment; filename={layout.codigo}.{extension}"
            },
        )
    # ======================================================
    # 🔵 PLANTILLA EXCEL (ASP / BANREGIO / ESPIRAL)
    # ======================================================
    elif tipo == "PLANTILLA_EXCEL":

        if not layout.ruta_plantilla:
            raise HTTPException(
                status_code=400, detail="Layout sin ruta_plantilla configurada"
            )

        wb = load_workbook(layout.ruta_plantilla)

        campos = (
            db.query(CatLayoutCampo)
            .filter(
                CatLayoutCampo.id_layout == id_layout,
                CatLayoutCampo.activo == True,
                CatLayoutCampo.tipo_campo.in_(["DETALLE_UI", "DETALLE_TXT"]),
            )
            .all()
        )

        if not campos:
            raise HTTPException(
                status_code=400, detail="Layout sin campos configurados"
            )

        # =====================================================
        # 🔹 SI EXISTE HOJA "Bancos", CONSTRUIR MAPA AUTOMÁTICO
        # =====================================================

        mapa_bancos = {}

        mapa_bancos = {k: v["nombre"] for k, v in MAPA_BANCOS.items()}

        # =====================================================
        # 🔹 AGRUPAR CAMPOS POR HOJA
        # =====================================================

        hojas = {}

        for campo in campos:
            hojas.setdefault(campo.hoja_excel, []).append(campo)

        # =====================================================
        # 🔹 LLENAR DATOS
        # =====================================================

        for hoja_nombre, campos_hoja in hojas.items():

            ws = wb[hoja_nombre]
            fila_base = campos_hoja[0].fila_inicio

            for index, registro in enumerate(registros):

                fila_actual = fila_base + index
                secuencia_actual = index + 1

                cuenta = registro.get("clabe", "")
                prefijo = str(cuenta)[:3]
                nombre_banco = mapa_bancos.get(prefijo, "")

                for campo in campos_hoja:

                    celda = ws[f"{campo.celda_excel}{fila_actual}"]

                    # 🔹 NO SOBREESCRIBIR FORMULAS
                    if celda.data_type == "f":
                        continue

                    # 🔹 SECUENCIA AUTOMATICA
                    # ✅ fix 2 — secuencia siempre se escribe primero
                    if campo.nombre_logico == "secuencia":
                        ws[f"{campo.celda_excel}{fila_actual}"] = int(secuencia_actual)
                        continue

                    # 🔹 RESOLVER BANCO POR PREFIJO (ESPIRAL)
                    elif campo.nombre_logico == "banco_nombre":
                        valor = nombre_banco

                    # 🔹 VALOR FIJO
                    elif campo.valor_fijo is not None:
                        valor = campo.valor_fijo

                    else:
                        valor = registro.get(campo.nombre_logico, "")

                    # 🔹 FORMATO NUMÉRICO
                    if campo.formato == "DECIMAL_2":
                        try:
                            valor = float(valor)
                        except:
                            valor = 0

                    ws[f"{campo.celda_excel}{fila_actual}"] = valor

        # =====================================================
        # 🔹 GUARDAR
        # =====================================================

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={layout.codigo}.xlsx"
            },
        )

    # ======================================================
    # 🔵 FIXED (SANTANDER)
    # ======================================================
    elif tipo == "FIXED":

        campos = (
            db.query(CatLayoutCampo)
            .filter(
                CatLayoutCampo.id_layout == id_layout, CatLayoutCampo.activo == True
            )
            .order_by(CatLayoutCampo.posicion_inicio)
            .all()
        )

        lineas = []

        for registro in registros:

            longitud_total = max(c.posicion_inicio + c.longitud - 1 for c in campos)

            buffer = [" "] * longitud_total

            for campo in campos:

                if campo.nombre_logico == "estado_cuenta":
                    valor = ""
                # ✅ Resolver banco_receptor desde CLABE
                elif campo.nombre_logico == "banco_receptor":
                    clabe = registro.get("clabe", "") or ""
                    prefijo = str(clabe)[:3]
                    valor = obtener_banco_santander(prefijo, db)
                elif campo.valor_fijo is not None:
                    valor = campo.valor_fijo
                else:
                    valor = registro.get(campo.nombre_logico, "") or "" ""

                if campo.formato == "NUMERIC":
                    if campo.nombre_logico in ["clabe"]:
                        valor = str(valor).ljust(campo.longitud)
                    else:
                        valor = str(valor).zfill(campo.longitud)

                elif campo.formato == "DECIMAL_2":
                    numero = float(valor or 0)
                    valor = str(int(round(numero, 2) * 100)).zfill(campo.longitud)

                elif campo.formato == "ALFANUM":
                    valor = str(valor).upper().ljust(campo.longitud)

                else:
                    valor = str(valor).ljust(campo.longitud)

                valor = valor[: campo.longitud]

                inicio = campo.posicion_inicio - 1
                fin = inicio + campo.longitud
                buffer[inicio:fin] = list(valor)

            lineas.append("".join(buffer))

        contenido = "\n".join(lineas)

        return StreamingResponse(
            BytesIO(contenido.encode("utf-8")),
            media_type="text/plain",
            headers={
                "Content-Disposition": f"attachment; filename={layout.codigo}.txt"
            },
        )

    # ======================================================
    # 🔴 ERROR SI NO COINCIDE NINGUNO
    # ======================================================
    else:
        raise HTTPException(
            status_code=400, detail=f"Tipo layout no soportado: {layout.tipo_layout}"
        )


@router.post("/generar_zip")
async def generar_zip(data: dict = Body(...), db: Session = Depends(get_db)):

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:

        for item in data["layouts"]:

            id_layout = item["id_layout"]
            registros = item["registros"]

            layout = db.get(CatLayout, id_layout)

            campos = (
                db.query(CatLayoutCampo)
                .filter(
                    CatLayoutCampo.id_layout == id_layout, CatLayoutCampo.activo == True
                )
                .all()
            )

            lineas = []

            for registro in registros:

                longitud_total = max(c.posicion_inicio + c.longitud - 1 for c in campos)

                buffer = [" "] * longitud_total

                for campo in campos:

                    valor = registro.get(campo.nombre_logico, "") or ""
                    valor = str(valor).ljust(campo.longitud)[: campo.longitud]

                    inicio = campo.posicion_inicio - 1
                    fin = inicio + campo.longitud
                    buffer[inicio:fin] = list(valor)

                lineas.append("".join(buffer).upper())

            contenido = "\n".join(lineas)

            zip_file.writestr(f"{layout.codigo}.txt", contenido)

    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=layouts.zip"},
    )


print("Bancos cargados:", len(MAPA_BANCOS))

cargar_bancos()
